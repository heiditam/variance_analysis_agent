"""Build the Summary + Drill-Down Excel workbook from account variance results."""

import os
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MATERIAL_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
BOLD = Font(bold=True)
TITLE_FONT = Font(bold=True, size=13)
HYPERLINK_FONT = Font(color="0563C1", underline="single")

SUMMARY_HEADERS = [
    "Dataset", "Account", "Period A Total", "Period B Total",
    "$ Change", "% Change", "Material", "Explanation",
]


def _pct_str(pct: Optional[float]) -> str:
    if pd.isna(pct):
        return "new"
    return f"{pct * 100:+.1f}%"


def _write_subheader(ws: Worksheet, row: int, text: str) -> None:
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUBHEADER_FILL
    ws.cell(row=row, column=1, value=text).font = BOLD


def _write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = BOLD
        cell.fill = HEADER_FILL


def short_explanation(text: str, max_len: int = 180) -> str:
    if not text:
        return ""
    first_sentence = text.split(". ")[0]
    if len(first_sentence) <= max_len:
        return first_sentence if first_sentence.endswith(".") else first_sentence + "."
    return text[:max_len].rstrip() + "..."


def _write_drilldown_block(
    ws: Worksheet,
    start_row: int,
    dataset: str,
    account: str,
    period_a: str,
    period_b: str,
    drilldown: dict,
    summary_row: int,
) -> int:
    row = start_row

    title = ws.cell(row=row, column=1, value=f"{account} — {dataset} — {period_a} vs {period_b}")
    title.font = TITLE_FONT
    row += 1

    back = ws.cell(row=row, column=1, value="↩ Back to Summary")
    back.hyperlink = f"#'Summary'!A{summary_row}"
    back.font = HYPERLINK_FONT
    row += 2

    _write_subheader(ws, row, "Category driver breakdown")
    row += 1
    cat_headers = ["Category", f"{period_a} Total", f"{period_b} Total", "$ Change", "Materiality Score", "Method"]
    _write_header_row(ws, row, cat_headers)
    row += 1
    for cat in drilldown.get("category_table", []):
        ws.cell(row=row, column=1, value=cat["category"])
        ws.cell(row=row, column=2, value=round(cat["total_a"], 2))
        ws.cell(row=row, column=3, value=round(cat["total_b"], 2))
        ws.cell(row=row, column=4, value=round(cat["abs_change"], 2))
        ws.cell(row=row, column=5, value=round(cat["materiality_score"], 3))
        ws.cell(row=row, column=6, value=cat["materiality_method"])
        row += 1
    row += 1

    concentration = drilldown.get("concentration")
    if concentration:
        _write_subheader(ws, row, f"Concentration within '{concentration['category']}'")
        row += 1
        con_headers = ["Contributor", f"{period_a} Total", f"{period_b} Total", "Delta", "Share of Change", "Role"]
        _write_header_row(ws, row, con_headers)
        row += 1
        for c in concentration.get("top_contributors", []):
            _write_contributor_row(ws, row, c, "Driver")
            row += 1
        for c in concentration.get("offsetting_contributors", []):
            _write_contributor_row(ws, row, c, "Offsetting")
            row += 1

        top_share = concentration.get("top_contributors_share")
        offsetting_share = concentration.get("offsetting_share")
        n = len(concentration.get("top_contributors", []))
        rollup = f"{n} contributor(s) account for {_pct_str(top_share)} of the change"
        if offsetting_share and offsetting_share != 0:
            rollup += f", partially offset by {_pct_str(offsetting_share)}"
        ws.cell(row=row, column=1, value=rollup).font = Font(italic=True)
        row += 1
        row += 1

    transactions = drilldown.get("transactions", [])
    if transactions:
        _write_subheader(ws, row, "Supporting transactions")
        row += 1
        _write_header_row(ws, row, ["Date", "Category", "Tag", "Amount"])
        row += 1
        for txn in transactions:
            ws.cell(row=row, column=1, value=str(txn["date"]))
            ws.cell(row=row, column=2, value=txn["category"])
            ws.cell(row=row, column=3, value=txn.get("tags", ""))
            ws.cell(row=row, column=4, value=round(txn["amount"], 2))
            row += 1
        row += 1

    full_explanation = drilldown.get("full_explanation", "")
    if full_explanation:
        _write_subheader(ws, row, "Full explanation")
        row += 1
        cell = ws.cell(row=row, column=1, value=full_explanation)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.row_dimensions[row].height = 45
        row += 1

    row += 2
    return row


def _write_contributor_row(ws: Worksheet, row: int, contributor: dict, role: str) -> None:
    ws.cell(row=row, column=1, value=str(contributor["key"]))
    ws.cell(row=row, column=2, value=round(contributor["total_a"], 2))
    ws.cell(row=row, column=3, value=round(contributor["total_b"], 2))
    ws.cell(row=row, column=4, value=round(contributor["delta"], 2))
    share = contributor.get("share_of_overall_change")
    ws.cell(row=row, column=5, value=_pct_str(share) if share is not None else "n/a")
    role_cell = ws.cell(row=row, column=6, value=role)
    if role == "Offsetting":
        role_cell.font = Font(italic=True)


def build_excel_report(
    account_rows: list[dict],
    drilldowns: dict,
    period_a: str,
    period_b: str,
    dataset_arg: str,
    out_dir: str = "reports",
) -> str:
    """Write the Summary + Drill-Down workbook. Returns the path written.

    account_rows: list of {dataset, account, total_a, total_b, abs_change,
      pct_change, is_material, explanation}.
    drilldowns: dict keyed by (dataset, account) -> {category_table, concentration,
      transactions, full_explanation} for material accounts.
    """
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_drilldown = wb.create_sheet("Drill-Down")

    # First pass: reserve a Summary row number for every account (header is row 1).
    summary_row_map = {
        (row["dataset"], row["account"]): i + 2 for i, row in enumerate(account_rows)
    }

    # Build the Drill-Down sheet first so we know each material account's anchor row.
    drilldown_row_map = {}
    d_row = 1
    for row in account_rows:
        key = (row["dataset"], row["account"])
        if not row["is_material"] or key not in drilldowns:
            continue
        drilldown_row_map[key] = d_row
        d_row = _write_drilldown_block(
            ws_drilldown, d_row, row["dataset"], row["account"],
            period_a, period_b, drilldowns[key], summary_row_map[key],
        )

    for col, width in enumerate([28, 16, 16, 14, 16, 12], start=1):
        ws_drilldown.column_dimensions[chr(64 + col)].width = width

    # Now write the Summary sheet, linking forward to the Drill-Down anchors.
    _write_header_row(ws_summary, 1, SUMMARY_HEADERS)
    for i, row in enumerate(account_rows):
        r = i + 2
        key = (row["dataset"], row["account"])
        ws_summary.cell(row=r, column=1, value=row["dataset"])
        account_cell = ws_summary.cell(row=r, column=2, value=row["account"])
        ws_summary.cell(row=r, column=3, value=round(row["total_a"], 2))
        ws_summary.cell(row=r, column=4, value=round(row["total_b"], 2))
        ws_summary.cell(row=r, column=5, value=round(row["abs_change"], 2))
        pct_cell = ws_summary.cell(row=r, column=6, value=_pct_str(row["pct_change"]))
        ws_summary.cell(row=r, column=7, value="Yes" if row["is_material"] else "No")
        ws_summary.cell(row=r, column=8, value=row.get("explanation", ""))

        if row["is_material"]:
            pct_cell.font = BOLD
            for col in range(1, 9):
                ws_summary.cell(row=r, column=col).fill = MATERIAL_FILL
            if key in drilldown_row_map:
                account_cell.hyperlink = f"#'Drill-Down'!A{drilldown_row_map[key]}"
                account_cell.font = HYPERLINK_FONT

    for col, width in enumerate([12, 14, 16, 16, 12, 12, 10, 60], start=1):
        ws_summary.column_dimensions[chr(64 + col)].width = width

    os.makedirs(out_dir, exist_ok=True)
    filename = f"variance_report_{dataset_arg}_{period_a}_vs_{period_b}.xlsx"
    out_path = os.path.join(out_dir, filename)
    wb.save(out_path)
    return out_path
