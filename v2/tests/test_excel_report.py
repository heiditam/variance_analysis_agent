import datetime

from openpyxl import load_workbook

from report.excel_report import build_excel_report, short_explanation


def _sample_data():
    account_rows = [
        {
            "dataset": "expenses", "account": "acct_1", "total_a": 100.0, "total_b": 200.0,
            "abs_change": 100.0, "pct_change": 1.0, "is_material": True,
            "explanation": "acct_1 increased 100%, driven by Food.",
        },
        {
            "dataset": "expenses", "account": "acct_2", "total_a": 100.0, "total_b": 103.0,
            "abs_change": 3.0, "pct_change": 0.03, "is_material": False, "explanation": "",
        },
    ]
    drilldowns = {
        ("expenses", "acct_1"): {
            "category_table": [
                {"category": "Food", "total_a": 50.0, "total_b": 150.0, "abs_change": 100.0,
                 "materiality_score": 0.9, "materiality_method": "zscore"},
            ],
            "concentration": {
                "category": "Food",
                "top_contributors": [
                    {"key": "tag_1", "total_a": 50.0, "total_b": 140.0, "delta": 90.0, "share_of_overall_change": 0.9},
                ],
                "offsetting_contributors": [],
                "top_contributors_share": 0.9,
                "offsetting_share": None,
            },
            "transactions": [
                {"date": datetime.date(2025, 11, 15), "category": "Food", "tags": "tag_1", "amount": 90.0},
            ],
            "full_explanation": "acct_1 increased 100%, driven by a spike in Food, with tag_1 accounting for 90% of the increase.",
        }
    }
    return account_rows, drilldowns


def test_build_excel_report_creates_both_sheets_with_expected_rows(tmp_path):
    account_rows, drilldowns = _sample_data()

    out_path = build_excel_report(account_rows, drilldowns, "2025-10", "2025-11", "expenses", out_dir=str(tmp_path))

    wb = load_workbook(out_path)
    assert "Summary" in wb.sheetnames
    assert "Drill-Down" in wb.sheetnames

    ws = wb["Summary"]
    assert ws.cell(row=1, column=2).value == "Account"
    assert ws.cell(row=2, column=2).value == "acct_1"
    assert ws.cell(row=2, column=7).value == "Yes"
    assert ws.cell(row=3, column=2).value == "acct_2"
    assert ws.cell(row=3, column=7).value == "No"


def test_material_row_hyperlink_resolves_and_non_material_row_has_none(tmp_path):
    account_rows, drilldowns = _sample_data()
    out_path = build_excel_report(account_rows, drilldowns, "2025-10", "2025-11", "expenses", out_dir=str(tmp_path))

    wb = load_workbook(out_path)
    ws_summary = wb["Summary"]
    ws_drilldown = wb["Drill-Down"]

    material_cell = ws_summary.cell(row=2, column=2)
    assert material_cell.hyperlink is not None
    target = material_cell.hyperlink.target
    assert target.startswith("#'Drill-Down'!A")
    target_row = int(target.split("!A")[1])
    assert ws_drilldown.cell(row=target_row, column=1).value.startswith("acct_1")

    non_material_cell = ws_summary.cell(row=3, column=2)
    assert non_material_cell.hyperlink is None
    assert not ws_summary.cell(row=3, column=8).value  # empty string round-trips as None via openpyxl


def test_short_explanation_truncates_long_text():
    long_text = "A" * 300
    result = short_explanation(long_text, max_len=50)
    assert len(result) <= 53  # max_len + "..."
    assert result.endswith("...")

    short_text = "Short sentence."
    assert short_explanation(short_text) == "Short sentence."

    assert short_explanation("") == ""
